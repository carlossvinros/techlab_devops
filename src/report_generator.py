import pandas as pd
import os
from typing import Dict, Any
import traceback
from openpyxl.utils import get_column_letter

def gerar_relatorio_excel(
    df_final_calculado: pd.DataFrame, 
    mapeamento_colunas: Dict[str, Any],
    caminho_output_dir: str, 
    nome_arquivo_output: str
) -> bool:
    """
    Gera o relatório final em formato Excel, com formatação de moeda
    e arredondamento para duas casas decimais nas colunas monetárias.
    """
    func_prefix = "[gerar_relatorio_excel]"
    print(f"\n--- Iniciando Geração do Relatório Excel com Formatação ---")

    if df_final_calculado is None or df_final_calculado.empty:
        print(f"{func_prefix} ERRO: DataFrame final está vazio ou não foi fornecido.")
        return False

    try:
        df_para_exportar = df_final_calculado.copy()

        colunas_monetarias = []
        coluna_salario = mapeamento_colunas.get("colaboradores", {}).get("nome_padronizado_custo")
        if coluna_salario and coluna_salario in df_para_exportar.columns:
            colunas_monetarias.append(coluna_salario)

        for nome_df, mapa in mapeamento_colunas.items():
            if nome_df != "colaboradores":
                custo_col_nome = mapa.get("nome_padronizado_custo")
                if custo_col_nome and custo_col_nome in df_para_exportar.columns and custo_col_nome not in colunas_monetarias:
                    colunas_monetarias.append(custo_col_nome)
        
        novas_colunas_subtotal_e_total = ["Centro_Custo_Ferramentas", "Centro_Custo_Beneficios", "Custo_Geral_Total"]
        for col in novas_colunas_subtotal_e_total:
            if col in df_para_exportar.columns and col not in colunas_monetarias:
                colunas_monetarias.append(col)
        
        print(f"{func_prefix} Colunas monetárias para formatação: {colunas_monetarias}")

        for col_monetaria in colunas_monetarias:
            if col_monetaria in df_para_exportar.columns:
                df_para_exportar[col_monetaria] = pd.to_numeric(df_para_exportar[col_monetaria], errors='coerce').round(2)
        
        print(f"{func_prefix} Dados monetários arredondados.")

        colunas_relatorio_ordenadas = ["CPF_Padronizado", "Nome_Padronizado"]
        if "Departamento" in df_para_exportar.columns:
            colunas_relatorio_ordenadas.append("Departamento")
        if coluna_salario and coluna_salario in df_para_exportar.columns:
            colunas_relatorio_ordenadas.append(coluna_salario)
        
        custos_ferramentas_ordenados = sorted([
            mapa['nome_padronizado_custo'] 
            for nome_df, mapa in mapeamento_colunas.items() 
            if nome_df in ["github", "google_workspace"] and mapa.get('nome_padronizado_custo') in df_para_exportar.columns
        ])
        colunas_relatorio_ordenadas.extend(custos_ferramentas_ordenados)
        if "Centro_Custo_Ferramentas" in df_para_exportar.columns:
            colunas_relatorio_ordenadas.append("Centro_Custo_Ferramentas")
        
        custos_beneficios_ordenados = sorted([
            mapa['nome_padronizado_custo'] 
            for nome_df, mapa in mapeamento_colunas.items() 
            if nome_df in ["gympass", "unimed"] and mapa.get('nome_padronizado_custo') in df_para_exportar.columns
        ])
        colunas_relatorio_ordenadas.extend(custos_beneficios_ordenados)
        if "Centro_Custo_Beneficios" in df_para_exportar.columns:
            colunas_relatorio_ordenadas.append("Centro_Custo_Beneficios")
        
        if "Custo_Geral_Total" in df_para_exportar.columns:
            colunas_relatorio_ordenadas.append("Custo_Geral_Total")
        
        colunas_finais_existentes = []
        for col in colunas_relatorio_ordenadas:
            if col in df_para_exportar.columns and col not in colunas_finais_existentes:
                colunas_finais_existentes.append(col)
        
        df_para_exportar = df_para_exportar[colunas_finais_existentes]
        print(f"{func_prefix} Ordem final das colunas: {list(df_para_exportar.columns)}")

        os.makedirs(caminho_output_dir, exist_ok=True)
        caminho_completo_output = os.path.join(caminho_output_dir, nome_arquivo_output)

        with pd.ExcelWriter(caminho_completo_output, engine='openpyxl') as writer:
            df_para_exportar.to_excel(writer, index=False, sheet_name='RateioDeCustos')
            
            worksheet = writer.sheets['RateioDeCustos']
            formato_moeda_brl = 'R$ #,##0.00'

            header_list = list(df_para_exportar.columns)
            for col_nome_monetaria in colunas_monetarias:
                if col_nome_monetaria in header_list:
                    col_idx_excel = header_list.index(col_nome_monetaria) + 1
                    letra_coluna = get_column_letter(col_idx_excel)
                    for i in range(2, worksheet.max_row + 1):
                        worksheet[f'{letra_coluna}{i}'].number_format = formato_moeda_brl
            print(f"{func_prefix} Formato de moeda aplicado às colunas monetárias.")

            for i, column_header in enumerate(header_list):
                column_letter = get_column_letter(i + 1)
                max_length = len(str(column_header))
                for row_idx in range(1, worksheet.max_row + 1):
                    cell_value = worksheet[f"{column_letter}{row_idx}"].value
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))
                
                adjusted_width = (max_length + 2) * 1.1 
                worksheet.column_dimensions[column_letter].width = adjusted_width
            print(f"{func_prefix} Largura das colunas ajustada.")

        print(f"{func_prefix} Relatório salvo e formatado com sucesso em: {caminho_completo_output}")
        print(f"--- Geração do Relatório Excel Concluída ---")
        return True

    except Exception as e:
        print(f"{func_prefix} ERRO ao gerar o relatório Excel formatado: {e}")
        traceback.print_exc()
        return False